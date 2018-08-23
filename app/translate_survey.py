# coding: utf-8

# TODO
"""
- Loop through and replace text in deserialised json instead of replacing values in string.
- Ignore any text between '<' and '>' and '{' and '}'.
- Sort out the loop for string replacement. There's got to be a better way of doing it than this!
- Find a way to remove all hardcoding. Can these be held globally in a properties file?
- Do we need to worry about ordering of JSON output?
- Add/amend other team/best practice stuff that's been missed out.
- Will this be run from command line or called from another script for now? Guard against different inputs.
"""
import json
import sys
import os
import re
from collections import OrderedDict
from openpyxl import load_workbook


def dumb_to_smart_quotes(string):
    """Takes a string and returns it with dumb quotes, single and double,
    replaced by smart quotes. Accounts for the possibility of HTML tags
    within the string.
    From https://gist.github.com/davidtheclark/5521432"""

    # Find dumb single quotes coming directly after letters or punctuation,
    # and replace them with right single quotes.
    string = re.sub(r"([\w.,?!;:\"\'])'", r'\1’', string)
    # Find any remaining dumb single quotes and replace them with
    # left single quotes.
    string = string.replace("'", '‘')
    # Reverse: Find any SMART quotes that have been (mistakenly) placed around HTML
    # attributes (following =) and replace them with dumb quotes.
    string = re.sub(r'=‘(.*?)’', r"='\1'", string)
    # Reverse: Find any SMART quotes that have been (mistakenly) placed around Jinja
    # attributes (following [) and replace them with dumb quotes.
    string = re.sub(r'\[‘(.*?)’', r"='\1'", string)
    return string.strip()


def translate_container(container, context, translations):
    if isinstance(container, dict):
        for key in ['description', 'label', 'title', 'question']:
            translate_value(container, context, key, translations)

    elif isinstance(container, list):
        for index, value in enumerate(container):
            source_key = (context, value)
            if source_key not in translations:
                print("No translation for container text '" + value + "'")
            else:
                container[index] = dumb_to_smart_quotes(translations[source_key])

    return container


def translate_value(container, context, key, translations):
    value = container.get(key)

    if value is not None and value != '':
        source_key = (context, value)

        if source_key not in translations:
            if 'format_household_name' not in value:
                print("No translation for text value '" + value + "' [" + context + "]")
        else:
            container[key] = dumb_to_smart_quotes(translations[source_key])


def translate_survey(survey_json, translations):
    translate_title_data(survey_json, translations)

    for section in survey_json['sections']:
        translate_container(section, section['id'], translations)

        for group in section['groups']:
            translate_container(group, group['id'], translations)

            for block in group['blocks']:
                if 'content' in block:
                    translate_content_block(block, block['id'], translations)

                translate_container(block, block['id'], translations)

                if 'questions' not in block:
                    translate_container(block, block['id'], translations)
                    translate_introduction_content(block, block, translations)
                    translate_calculated_summary(block, translations)
                    continue

                for question in block['questions']:
                    translate_container(question, question['id'], translations)
                    translate_validation_text(question, question['id'], translations)
                    translate_guidance_text(question, question['id'], translations)
                    translate_titles_text(question, question['id'], translations)
                    translate_definitions_text(question, question['id'], translations)

                    for answer in question['answers']:
                        translate_container(answer, answer['id'], translations)
                        translate_guidance_text(answer, answer['id'], translations)
                        translate_options_text(answer, answer['id'], translations)
                        translate_validation_text(answer, answer['id'], translations)

    return survey_json


def translate_title_data(container, translations):
    translate_value(container, 'title', 'title', translations)
    translate_value(container, 'description', 'description', translations)
    translate_value(container, 'legal_basis', 'legal_basis', translations)


def translate_content_block(container, context, translations):
    if 'content' in container:
        for content in container['content']:
            translate_container(content, context, translations)

            if 'list' in content:
                content['list'] = translate_container(content['list'], context, translations)

    return container


def translate_validation_text(container, context, translations):
    if 'validation' in container:
        for key, value in container['validation']['messages'].items():
            source_key = (context + ' [validation message]', value)
            if source_key not in translations:
                print("No translation for text '" + value + "'")
            else:
                container['validation']['messages'][key] = dumb_to_smart_quotes(translations[source_key])

    return container


def translate_guidance_text(container, context, translations):
    if 'guidance' in container:
        for guide in container['guidance']:
            translate_container(container['guidance'], context + ' [question guidance]', translations)

            if 'hide_guidance' in guide:
                translate_show_hide_text(container['guidance'], context + ' [question guidance]', translations)

        for guidance in container['guidance']['content']:
            translate_container(guidance, context + ' [question guidance]', translations)

            if 'list' in guidance:
                guidance['list'] = translate_container(guidance['list'], context + ' [question guidance]', translations)

    return container


def translate_show_hide_text(container, context, translations):
    for value in container:
        if 'show_guidance' in value:
            translate_value(container, context, value, translations)
        elif 'hide_guidance' in value:
            translate_value(container, context, value, translations)
    return container


def translate_titles_text(container, context, translations):
    if 'titles' in container:
        for titles in container['titles']:
            if isinstance(titles, dict):
                for key in ['description', 'label', 'title', 'question', 'value']:
                    translate_value(titles, context, key, translations)

    return container


def translate_options_text(container, context, translations):
    if 'options' in container:
        for options in container['options']:
            translate_container(options, context, translations)

    return container


def translate_primary_content_text(container, context, translations):
    if 'primary_content' in container:
        primary_content_text = container['primary_content'][0]

        for primary_content in primary_content_text:
            if 'content' in primary_content:
                content = primary_content_text['content']

                for values in content:
                    if 'list' in values:
                        values['list'] = translate_container(values['list'], context, translations)
                    else:
                        translate_container(values, context, translations)

            else:
                translate_container(primary_content_text, context, translations)

    return container


def translate_preview_content_text(container, context, translations):
    if 'preview_content' in container:
        preview_content_text = container['preview_content']

        for preview_content in preview_content_text:
            if 'content' in preview_content:
                content = preview_content_text['content']

                for values in content:
                    if 'list' in values:
                        values['list'] = translate_container(values['list'], context, translations)
                    translate_container(values, context, translations)

            elif 'questions' in preview_content:
                questions = preview_content_text['questions']

                for values in questions:
                    if 'content' in values:
                        content = values['content']

                        for value in content:
                            if 'list' in value:
                                value['list'] = translate_container(value['list'], context, translations)
                            translate_container(value, context, translations)

                    translate_container(values, context, translations)

            else:
                translate_container(preview_content_text, context, translations)

    return container


def translate_secondary_content_text(container, context, translations):
    if 'secondary_content' in container:
        secondary_content_text = container['secondary_content'][0]

        for secondary_content in secondary_content_text:
            if 'content' in secondary_content:
                content = secondary_content_text['content']

                for values in content:
                    if 'list' in values:
                        values['list'] = translate_container(values['list'], context, translations)
                    translate_container(values, context, translations)

            else:
                translate_container(secondary_content_text, context, translations)

    return container


def translate_introduction_content(container, context, translations):
    if 'primary_content' in context:
        translate_primary_content_text(container, context['primary_content'][0]['id'], translations)
    if 'preview_content' in context:
        translate_preview_content_text(container, context['preview_content']['id'], translations)
    if 'secondary_content' in context:
        translate_secondary_content_text(container, context['secondary_content'][0]['id'], translations)


def translate_calculated_summary(container, translations):
    if container['type'] == 'CalculatedSummary':
        translate_titles_text(container, container['id'], translations)
        translate_titles_text(container['calculation'], container['id'], translations)


def translate_definitions_text(container, context, translations):
    if 'definitions' in container:
        for definition in container['definitions']:
            translate_container(definition, context, translations)
            translate_content_block(definition, context, translations)

    return container


def load_translations(input_file):
    wb = load_workbook(input_file)
    sheet = wb.get_sheet_by_name('Sheet')

    translations = {}
    for row in sheet.iter_rows(row_offset=1, min_col=1, max_col=3):
        source_key = (row[0].value, row[1].value)
        translated_text = row[2].value
        if source_key is not None:
            if translated_text is not None:
                translations[source_key] = translated_text

    return translations


def deserialise_json(json_file_to_deserialise):
    with open(json_file_to_deserialise, 'r', encoding="utf8") as json_data:
        try:
            data = json.load(json_data, object_pairs_hook=OrderedDict)
            return data

        except ValueError:
            print("Error decoding JSON. Please ensure file is in valid JSON format.")
            return None


def strip_directory_and_extension(file):
    file_basename = os.path.basename(file)
    file_name = os.path.splitext(file_basename)[0]

    return file_name


def create_output_file_name_with_directory(output_directory, input_file):
    file_name = strip_directory_and_extension(json_file)
    file_name_with_extension = file_name + '.json'
    file_name_with_directory = os.path.join(output_directory, file_name_with_extension)

    return file_name_with_directory


def save_translated_json(translated_json, output_file_name):
    output = json.dumps(translated_json, indent=4, ensure_ascii=False, separators=(',', ': '))
    with open(output_file_name, "w", encoding="utf8") as target_file:
        target_file.writelines(output)


def command_line_handler(json_file, input_file, output_directory):
    survey_json = deserialise_json(json_file)
    if survey_json is None:
        exit(1)

    translations = load_translations(input_file)
    if translations is None:
        exit(1)

    print()
    translated_json = translate_survey(survey_json, translations)
    if translated_json is None:
        exit(1)

    output_file_name = create_output_file_name_with_directory(output_directory, json_file)
    save_translated_json(translated_json, output_file_name)

    BOLD = '\033[1m'
    GREEN = '\033[92m'
    END = '\033[0m'
    print()
    print(BOLD + GREEN + 'SUCCESS' + END + ' - Translated JSON saved at ' + output_file_name)
    print()
    exit(0)


if __name__ == '__main__':

    json_file = sys.argv[1]
    input_file = sys.argv[2]
    output_directory = sys.argv[3]

    command_line_handler(json_file, input_file, output_directory)
