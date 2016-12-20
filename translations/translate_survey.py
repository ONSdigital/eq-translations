# TODO
"""
- Loop through and replace text in deserialised json instead of replacing values in string.
- Ignore any text between '<' and '>' and '{' and '}'.
- Sort out the loop for string replacement. There's got to be a better way of doing it than this!
- Find a way to remove all hardcoding. Can these be held globally in a properties file?
- Do we need to worry about ordering of JSON output?
- Add/amend other team/best practice stuff that's been missed out.
- Will this be run from command line or called from another script for now? Guard against different inputs.
"""

import json
import sys
import os
import re
from collections import OrderedDict
from openpyxl import Workbook, load_workbook


def dumb_to_smart_quotes(string):
    """Takes a string and returns it with dumb quotes, single and double,
    replaced by smart quotes. Accounts for the possibility of HTML tags
    within the string.
    From https://gist.github.com/davidtheclark/5521432"""

    # Find dumb single quotes coming directly after letters or punctuation,
    # and replace them with right single quotes.
    string = re.sub(r"([a-zA-Z0-9.,?!;:\"\'])'", r'\1’', string)
    # Find any remaining dumb single quotes and replace them with
    # left single quotes.
    string = string.replace("'", '‘')
    # Reverse: Find any SMART quotes that have been (mistakenly) placed around HTML
    # attributes (following =) and replace them with dumb quotes.
    string = re.sub(r'=‘(.*?)’', r"='\1'", string)
    return string.strip()


def translate_container(container, context, translations):
    if isinstance(container, dict):
        for key in ['description', 'label', 'title']:
            value = container.get(key)

            if value is not None and value != '':
                source_key = (context, value)
                if source_key not in translations:
                    if 'format_household_name' not in value:
                        print("No translation for text '" + value + "' [" + context + "]")
                else:
                    container[key] = dumb_to_smart_quotes(translations[source_key])

    elif isinstance(container, list):
        for index, value in enumerate(container):
            source_key = (context, value)
            if source_key not in translations:
                print("No translation for text '" + value + "'")
            else:
                container[index] = dumb_to_smart_quotes(translations[source_key])

    return container


def translate_survey(survey_json, translations):
    for group in survey_json['groups']:
        translate_container(group, group['id'], translations)

        for block in group['blocks']:

            for section in block['sections']:
                translate_container(section, section['id'], translations)

                for question in section['questions']:
                    translate_container(question, question['id'], translations)
                    translate_validation_text(question, question['id'], translations)
                    translate_guidance_text(question, question['id'], translations)

                    for answer in question['answers']:
                        translate_container(answer, answer['id'], translations)
                        translate_guidance_text(answer, answer['id'], translations)
                        translate_options_text(answer, answer['id'], translations)
                        translate_validation_text(answer, answer['id'], translations)

    return survey_json


def translate_validation_text(container, context, translations):
    if 'validation' in container:
        for key, value in container['validation']['messages'].items():
            source_key = (context + ' [validation message]', value)
            if source_key not in translations:
                print("No translation for text '" + value + "'")
            else:
                container['validation']['messages'][key] = dumb_to_smart_quotes(translations[source_key])

    return container


def translate_guidance_text(container, context, translations):
    if 'guidance' in container:
        guidance_text = container['guidance']

        if isinstance(guidance_text, str):
            source_key = (context + ' [answer guidance]', guidance_text)
            if source_key not in translations:
                print("No translation for text '" + guidance_text + "'")
            else:
                container['guidance'] = dumb_to_smart_quotes(translations[source_key])
        else:
            for guidance in container['guidance']:
                translate_container(guidance, context + ' [question guidance]', translations)

                if 'list' in guidance:
                    guidance['list'] = translate_container(guidance['list'], context + ' [question guidance]', translations)

    return container


def translate_options_text(container, context, translations):
    if 'options' in container:
        for options in container['options']:
            options = translate_container(options, context, translations)

            if 'other' in options:
                options['other'] = translate_container(options['other'], context, translations)

    return container


def load_translations(input_file):
    wb = Workbook()
    wb = load_workbook(input_file)
    sheet = wb.get_sheet_by_name('Sheet')

    translations = {}
    for row in sheet.iter_rows(row_offset=2, min_col=1, max_col=3):
        source_key = (row[0].value, row[1].value)
        translated_text = row[2].value
        if source_key is not None:
            if translated_text is not None:
                translations[source_key] = translated_text

    return translations


def deserialise_json(json_file_to_deserialise):
    with open(json_file_to_deserialise, 'r', encoding="utf8") as json_data:
        try:
            data = json.load(json_data, object_pairs_hook=OrderedDict)
            return data

        except ValueError:
            print("Error decoding JSON. Please ensure file is in valid JSON format.")
            return None


def strip_directory_and_extension(file):
    file_basename = os.path.basename(file)
    file_name = os.path.splitext(file_basename)[0]

    return file_name


def create_output_file_name_with_directory(output_directory, input_file):
    file_name = strip_directory_and_extension(json_file)
    file_name_with_extension = file_name + '_cy.json'
    file_name_with_directory = os.path.join(output_directory, file_name_with_extension)

    return file_name_with_directory

def save_translated_json(translated_json, output_file_name):
    output = json.dumps(translated_json, indent=4, ensure_ascii=False, separators=(',', ': '))
    with open(output_file_name, "w", encoding="utf8") as target_file:
        target_file.writelines(output)


def command_line_handler(json_file, input_file, output_directory):
    survey_json = deserialise_json(json_file)
    if survey_json is None:
        exit(1)

    translations = load_translations(input_file)
    if translations is None:
        exit(1)

    print()
    translated_json = translate_survey(survey_json, translations)
    if translated_json is None:
        exit(1)

    output_file_name = create_output_file_name_with_directory(output_directory, json_file)
    save_translated_json(translated_json, output_file_name)

    BOLD = '\033[1m'
    GREEN = '\033[92m'
    END = '\033[0m'
    print()
    print(BOLD + GREEN + 'SUCCESS' + END + ' - Translated JSON saved at ' + output_file_name)
    print()
    exit(0)

if __name__ == '__main__':

    json_file = sys.argv[1]
    input_file = sys.argv[2]
    output_directory = sys.argv[3]

    command_line_handler(json_file, input_file, output_directory)
